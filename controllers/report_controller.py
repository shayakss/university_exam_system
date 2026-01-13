"""
Report Controller
Generates Excel and PDF reports
"""
from database.db_manager import db
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyQt5.QtWidgets import QTextEdit
from PyQt5.QtGui import QTextDocument, QPageSize
from PyQt5.QtPrintSupport import QPrinter

class ReportController:
    """Controller for generating reports"""
    
    def generate_excel_report(self, title: str, headers: List[str], data: List[List], filename: str) -> Tuple[bool, str]:
        """Generate a generic Excel report"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center")
            
            # Headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                cell.border = Border(bottom=Side(style='thin'))
            
            # Data
            for row_idx, row_data in enumerate(data, 3):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save
            save_dir = os.path.join(os.getcwd(), "reports")
            os.makedirs(save_dir, exist_ok=True)
            full_path = os.path.join(save_dir, filename)
            wb.save(full_path)
            
            return True, full_path
            
        except Exception as e:
            return False, f"Error generating Excel: {str(e)}"

    def generate_pdf_report(self, html_content: str, filename: str) -> Tuple[bool, str]:
        """Generate a PDF report from HTML"""
        try:
            document = QTextDocument()
            document.setHtml(html_content)
            
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            
            save_dir = os.path.join(os.getcwd(), "reports")
            os.makedirs(save_dir, exist_ok=True)
            full_path = os.path.join(save_dir, filename)
            
            printer.setOutputFileName(full_path)
            printer.setPageSize(QPrinter.A4)
            
            document.print_(printer)
            
            return True, full_path
            
        except Exception as e:
            return False, f"Error generating PDF: {str(e)}"

    def get_student_performance_data(self, department_id: int = None) -> Tuple[List[str], List[List]]:
        """Get data for student performance report"""
        query = """
            SELECT s.roll_number, s.name, d.department_name, s.semester,
                   r.cgpa, r.percentage, r.overall_grade, r.status
            FROM students s
            JOIN departments d ON s.department_id = d.department_id
            LEFT JOIN results r ON s.student_id = r.student_id
            WHERE s.is_active = 1
        """
        params = []
        if department_id:
            query += " AND s.department_id = ?"
            params.append(department_id)
            
        query += " ORDER BY s.roll_number"
        
        results = db.execute_query(query, tuple(params))
        
        headers = ["Roll No", "Name", "Department", "Semester", "CGPA", "Percentage", "Grade", "Status"]
        data = []
        for r in results:
            data.append([
                r['roll_number'], r['name'], r['department_name'], r['semester'],
                r['cgpa'] or 0.0, r['percentage'] or 0.0, r['overall_grade'] or '-', r['status'] or '-'
            ])
            
        return headers, data

    def get_attendance_data(self, department_id: int = None) -> Tuple[List[str], List[List]]:
        """Get data for attendance report"""
        query = """
            SELECT s.roll_number, s.name, d.department_name,
                   COUNT(sa.attendance_id) as total_days,
                   SUM(CASE WHEN sa.status = 'Present' THEN 1 ELSE 0 END) as present_days,
                   SUM(CASE WHEN sa.status = 'Absent' THEN 1 ELSE 0 END) as absent_days,
                   SUM(CASE WHEN sa.status = 'Late' THEN 1 ELSE 0 END) as late_days
            FROM students s
            JOIN departments d ON s.department_id = d.department_id
            LEFT JOIN student_attendance sa ON s.student_id = sa.student_id
            WHERE s.is_active = 1
        """
        params = []
        if department_id:
            query += " AND s.department_id = ?"
            params.append(department_id)
            
        query += " GROUP BY s.student_id ORDER BY s.roll_number"
        
        results = db.execute_query(query, tuple(params))
        
        headers = ["Roll No", "Name", "Department", "Total Days", "Present", "Absent", "Late", "Percentage"]
        data = []
        for r in results:
            total = r['total_days']
            present = r['present_days'] + r['late_days'] # Late counts as present usually
            pct = (present / total * 100) if total > 0 else 0
            
            data.append([
                r['roll_number'], r['name'], r['department_name'],
                total, r['present_days'], r['absent_days'], r['late_days'], f"{pct:.1f}%"
            ])
            
        return headers, data

# Global instance
report_controller = ReportController()
