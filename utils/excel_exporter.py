"""
Excel Exporter - Export data to Excel files
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict
import config


class ExcelExporter:
    """Export data to Excel files"""
    
    def export_student_list(self, students: List[dict], output_path: str) -> bool:
        """Export student list to Excel"""
        try:
            # Prepare data
            data = []
            for student in students:
                data.append({
                    'Roll Number': student['roll_number'],
                    'Name': student['name'],
                    'Department': student.get('department_name', ''),
                    'Semester': student['semester'],
                    'Gender': student.get('gender', ''),
                    'Date of Birth': student.get('date_of_birth', ''),
                    'Email': student.get('email', ''),
                    'Phone': student.get('phone', '')
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Students', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['Students']
                self._format_worksheet(worksheet)
            
            print(f"✓ Student list exported: {output_path}")
            return True
        
        except Exception as e:
            print(f"✗ Excel export error: {e}")
            return False
    
    def export_marks_report(self, marks: List[dict], output_path: str, 
                           title: str = "Marks Report") -> bool:
        """Export marks report to Excel"""
        try:
            # Prepare data
            data = []
            for mark in marks:
                data.append({
                    'Roll Number': mark.get('roll_number', ''),
                    'Student Name': mark.get('student_name', ''),
                    'Course Code': mark.get('course_code', ''),
                    'Course Name': mark.get('course_name', ''),
                    'Max Marks': mark['max_marks'],
                    'Marks Obtained': mark['marks_obtained'],
                    'Grade': mark['grade'],
                    'Status': mark['status']
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Marks', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['Marks']
                self._format_worksheet(worksheet)
            
            print(f"✓ Marks report exported: {output_path}")
            return True
        
        except Exception as e:
            print(f"✗ Excel export error: {e}")
            return False
    
    def export_results_report(self, results: List[dict], output_path: str) -> bool:
        """Export results report to Excel"""
        try:
            # Prepare data
            data = []
            for result in results:
                data.append({
                    'Roll Number': result.get('roll_number', ''),
                    'Student Name': result.get('student_name', ''),
                    'Department': result.get('department_name', ''),
                    'Semester': result['semester'],
                    'Total Marks': result['total_marks'],
                    'Marks Obtained': result['marks_obtained'],
                    'Percentage': f"{result['percentage']}%",
                    'SGPA': result['sgpa'],
                    'CGPA': result['cgpa'],
                    'Grade': result['overall_grade'],
                    'Status': result['status'],
                    'Rank': result.get('rank', '')
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Results', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['Results']
                self._format_worksheet(worksheet)
            
            print(f"✓ Results report exported: {output_path}")
            return True
        
        except Exception as e:
            print(f"✗ Excel export error: {e}")
            return False
    
    def export_detailed_marksheet(self, result_data: dict, output_path: str) -> bool:
        """Export detailed marksheet for a single student"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Marksheet"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = config.UNIVERSITY_NAME
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:G2')
            ws['A2'] = "STUDENT MARKSHEET"
            ws['A2'].font = Font(size=14, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Student info
            student = result_data['student']
            row = 4
            ws[f'A{row}'] = "Roll Number:"
            ws[f'B{row}'] = student['roll_number']
            ws[f'D{row}'] = "Name:"
            ws[f'E{row}'] = student['name']
            
            row += 1
            ws[f'A{row}'] = "Department:"
            ws[f'B{row}'] = student['department_name']
            ws[f'D{row}'] = "Semester:"
            ws[f'E{row}'] = result_data['semester']
            
            # Marks table header
            row += 2
            headers = ['S.No', 'Course Code', 'Course Name', 'Max Marks', 'Marks Obtained', 'Grade', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Marks data
            for idx, mark in enumerate(result_data['marks'], 1):
                row += 1
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=mark['course_code'])
                ws.cell(row=row, column=3, value=mark['course_name'])
                ws.cell(row=row, column=4, value=mark['max_marks'])
                ws.cell(row=row, column=5, value=mark['marks_obtained'])
                ws.cell(row=row, column=6, value=mark['grade'])
                ws.cell(row=row, column=7, value=mark['status'])
            
            # Total row
            row += 1
            ws.cell(row=row, column=3, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=4, value=result_data['total_marks']).font = Font(bold=True)
            ws.cell(row=row, column=5, value=result_data['marks_obtained']).font = Font(bold=True)
            
            # Result summary
            row += 2
            ws[f'A{row}'] = "Percentage:"
            ws[f'B{row}'] = f"{result_data['percentage']}%"
            ws[f'D{row}'] = "SGPA:"
            ws[f'E{row}'] = result_data['sgpa']
            
            row += 1
            ws[f'A{row}'] = "CGPA:"
            ws[f'B{row}'] = result_data['cgpa']
            ws[f'D{row}'] = "Overall Grade:"
            ws[f'E{row}'] = result_data['overall_grade']
            
            row += 1
            ws[f'A{row}'] = "Result:"
            ws[f'B{row}'] = result_data['status']
            ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'B{row}'].fill = PatternFill(
                start_color="27AE60" if result_data['status'] == 'Pass' else "E74C3C",
                end_color="27AE60" if result_data['status'] == 'Pass' else "E74C3C",
                fill_type="solid"
            )
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            print(f"✓ Detailed marksheet exported: {output_path}")
            return True
        
        except Exception as e:
            print(f"✗ Excel export error: {e}")
            return False
    
    def _format_worksheet(self, worksheet):
        """Apply formatting to worksheet"""
        # Header row formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


# Global Excel exporter instance
excel_exporter = ExcelExporter()
